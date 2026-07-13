import os
import re
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from PIL import Image
from dotenv import load_dotenv

# Variables globales para los mensajes de la terminal
CORTE_LINEA = "=================================================================="

print("=============================================")
print("   MOTOR HÍBRIDO DE EXTRACCIÓN DE FACTURAS   ")
print("=============================================\n")

# 1. MENÚ DE OPCIONES DE ENTRADA CORREGIDO
print("¿Qué deseas procesar hoy?")
print("1. Una o pocas facturas puntuales (factura/ticket)")
print("2. Un lote completo (Carpeta de imágenes)")
opcion_modo = input("Elige una opción (1 o 2): ").strip()

root = tk.Tk()
root.withdraw()
root.attributes('-topmost', True)

rutas_imagenes = []

if opcion_modo == "1":
    rutas = filedialog.askopenfilenames(
        title="Selecciona la(s) factura(s) o ticket(s)",
        filetypes=[("Imágenes", "*.jpg *.jpeg *.png *.bmp *.webp *.JPG *.JPEG")]
    )
    if rutas: rutas_imagenes = list(rutas)
elif opcion_modo == "2":
    carpeta = filedialog.askdirectory(title="Selecciona la carpeta con las facturas")
    if carpeta:
        extensiones_validas = ('.jpg', '.jpeg', '.png', '.bmp', '.webp', '.JPG', '.JPEG')
        rutas_imagenes = [os.path.join(carpeta, f) for f in os.listdir(carpeta) if f.lower().endswith(extensiones_validas)]
else:
    print("\n❌ Opción no válida. Proceso cancelado.")
    exit()

if not rutas_imagenes:
    print("\n❌ No se seleccionaron archivos. Proceso cancelado.")
    exit()

# Cargar configuración desde accesoimagen.env para la API de Gemini
load_dotenv("accesoimagen.env")
api_key = os.getenv("GEMINI_API_KEY")

archivo_borrador_excel = "vista_previa_facturas.xlsx"
datos_extraidos = []

# ==========================================
# MOTOR INTERNO 1: ESTRATEGIA CLOUD (GEMINI)
# ==========================================
def intentar_extraccion_cloud(ruta_img):
    from google import genai
    client = genai.Client(api_key=api_key)
    
    prompt = """
    Analiza esta imagen de factura o ticket. Extrae exclusivamente dos datos y devuélvelos exactamente en este formato, sin comentarios, introducciones ni bloques markdown:
    FECHA: dd/mm/yyyy
    TOTAL: numero_limpio

    Reglas:
    1. En FECHA, unifica usando barras (/). Si tiene guiones o espacios, conviértela obligatoriamente a 'dd/mm/yyyy' con año de 4 dígitos. Si no existe, pon X.
    2. En TOTAL, localiza el monto final definitivo real (Gran Total). Devuelve solo el número usando punto para decimales (ejemplo: 5165.00). No incluyas letras ni símbolos de moneda.
    """
    
    img = Image.open(ruta_img)
    response = client.models.generate_content(
        model='gemini-2.5-flash',
        contents=[prompt, img]
    )
    
    texto_ia = response.text.strip()
    match_f = re.search(r'(?i)FECHA:\s*([^\n]+)', texto_ia)
    match_t = re.search(r'(?i)TOTAL:\s*([0-9.]+)', texto_ia)
    
    fecha = match_f.group(1).strip() if match_f else "X"
    monto = float(match_t.group(1).strip()) if match_t else "X"
    return fecha, monto

# ==========================================
# MOTOR INTERNO 2: CONTINGENCIA LOCAL (EASYOCR)
# ==========================================
def intentar_extraccion_local(ruta_img):
    import easyocr
    # Inicialización tardía para no consumir memoria RAM si hay internet
    reader = easyocr.Reader(['es'], gpu=False)
    
    img_original = Image.open(ruta_img)
    ancho, alto = img_original.size
    
    # REGLA GEOMÉTRICA: Recortar el último cuadrante (Esquina Inferior Derecha)
    # Box: (izquierda, arriba, derecha, abajo)
    recorte_box = (int(ancho * 0.50), int(alto * 0.70), ancho, alto)
    img_recortada = img_original.crop(recorte_box)
    
    # Guardar borrador temporal del pedazo para pasarlo a EasyOCR
    ruta_temp_crop = "temp_crop_failover.jpg"
    img_recortada.save(ruta_temp_crop)
    
    # Escanear solo el pedazo de manuscrito aislado
    resultados = reader.readtext(ruta_temp_crop, detail=0)
    
    # Limpieza inmediata del archivo temporal en disco
    if os.path.exists(ruta_temp_crop):
        os.remove(ruta_temp_crop)
        
    texto_recorte = " ".join(resultados)
    
    # El francotirador busca números estándar con decimales escritos por ti
    encontrados = re.findall(r'\b\d+[\.,]\d{1,2}\b', texto_recorte)
    monto_local = "X"
    
    if encontrados:
        m_norm = encontrados[-1].replace(',', '.') # Toma el último por seguridad
        try:
            monto_local = float(m_norm)
        except ValueError:
            pass
    else:
        # Contingencia si el manuscrito no se leyó con los decimales
        numeros_enteros = re.findall(r'\b\d+\b', texto_recorte)
        if numeros_enteros:
            try:
                monto_local = float(max(numeros_enteros, key=len))
            except ValueError:
                pass

    # El modo contingencia local asume Fecha "X" por diseño técnico
    return "X", monto_local

# ==========================================
# FLUJO PRINCIPAL CONTROLADO (FAILOVER)
# ==========================================
try:
    print(f"\n⚙️ Iniciando procesamiento de {len(rutas_imagenes)} archivo(s)...")
    
    for idx, ruta in enumerate(rutas_imagenes, 1):
        nombre_archivo = os.path.basename(ruta)
        nombre_sin_extension = os.path.splitext(nombre_archivo)[0]
        
        print(f"\n[{idx}/{len(rutas_imagenes)}] Analizando: {nombre_archivo}")
        
        f_fecha = "X"
        f_monto = "X"
        
        # INTENTO 1: Intentar por defecto usar la nube inteligente (Gemini)
        try:
            if not api_key:
                raise ConnectionError("No hay API Key configurada.")
                
            f_fecha, f_monto = intentar_extraccion_cloud(ruta)
            print("   ✅ Procesado exitosamente mediante IA Cloud.")
            
        except Exception:
            # INTENTO 2: CONMUTACIÓN POR ERROR AUTOMÁTICA (MODO LOCAL)
            print("\n" + CORTE_LINEA)
            print("⚠️  [CONEXIÓN FALLIDA / ERROR CLOUD]")
            print("    Activando Modo Contingencia Local...")
            print("    -> Buscando manuscrito en esquina inferior derecha...")
            print(CORTE_LINEA + "\n")
            
            try:
                f_fecha, f_monto = intentar_extraccion_local(ruta)
                print("   ✅ Procesado exitosamente mediante Contingencia Local.")
            except Exception as e_local:
                print(f"   ❌ Error crítico en el motor local: {e_local}")
        
        # Registrar en la lista maestra con el formato de columnas exacto solicitado
        datos_extraidos.append({
            "ARCHIVO ORIGEN": nombre_sin_extension,
            "FECHA": f_fecha,
            "MONTO TOTAL": f_monto
        })

    # ==========================================
    # CONSTRUCCIÓN DE LA TABLA LIMPIA EXCEL
    # ==========================================
    print("\n📊 Generando matriz de vista previa en Excel...")
    df = pd.DataFrame(datos_extraidos)
    df.to_excel(archivo_borrador_excel, index=False)
    
    libro = load_workbook(archivo_borrador_excel)
    hoja = libro.active
    
    fuente_headers = Font(name="Segoe UI", size=11, bold=True)
    fuente_cuerpo = Font(name="Segoe UI", size=10)
    align_centro = Alignment(horizontal="center", vertical="center")
    align_izq = Alignment(horizontal="left", vertical="center")
    borde_fino = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    if hoja.views.sheetView:
        hoja.views.sheetView[0].showGridLines = True

    for num_fila, fila in enumerate(hoja.iter_rows(min_row=1), start=1):
        es_header = (num_fila == 1)
        for num_col, celda in enumerate(fila, start=1):
            celda.border = borde_fino
            
            if es_header:
                celda.font = fuente_headers
                celda.alignment = align_centro
            else:
                celda.font = fuente_cuerpo
                # Columna 1 (Archivo) a la izquierda, demás centradas
                if num_col == 1:
                    celda.alignment = align_izq
                else:
                    celda.alignment = align_centro
                
                # Regla Estricta: Formato numérico limpio contable (0.00) en Columna 3
                if num_col == 3 and celda.value != "X":
                    try:
                        celda.value = float(celda.value)
                        celda.number_format = '0.00'
                    except (ValueError, TypeError):
                        pass

    # Dimensionamiento fijo y limpio para evitar deformaciones visuales
    for col in hoja.columns:
        col_letter = col[0].column_letter
        if col_letter == 'A':
            hoja.column_dimensions[col_letter].width = 40
        else:
            hoja.column_dimensions[col_letter].width = 20
        
    libro.save(archivo_borrador_excel)
    
    # Abrir vista previa directamente en VS Code
    os.system(f'code {archivo_borrador_excel}')
    
    print("\n" + CORTE_LINEA)
    print(f"📊 ¡Proceso completo! Revisa la pestaña '{archivo_borrador_excel}' en VS Code.")
    print("   El formato de los datos es plano, unificado y original.")
    print(CORTE_LINEA)
    
    input("\n▶ Presiona [ENTER] aquí en la terminal cuando estés listo para exportar...")
    
    desea_guardar = input("❓ ¿Deseas guardar este lote de datos definitivamente? (S/N): ").strip().lower()
    
    if desea_guardar == 's':
        nombre_final = input("📝 Nombre definitivo para tu Excel (sin extensión): ").strip()
        nombre_archivo_final = f"{nombre_final}.xlsx"
        
        libro.save(nombre_archivo_final)
        print(f"\n✅ ¡Éxito contable! El lote limpio se guardó en: {nombre_archivo_final}")
    else:
        print("\n👋 Proceso cerrado. Datos temporales removidos.")

except Exception as e:
    print(f"\n❌ Ocurrió un error inesperado en el sistema general: {e}")

finally:
    # Garantizar la limpieza de disco eliminando el borrador contable pase lo que pase
    if os.path.exists(archivo_borrador_excel):
        try:
            os.remove(archivo_borrador_excel)
        except Exception:
            pass